import re
from datetime import datetime
from urllib.parse import unquote, urlparse, parse_qs

import openpyxl
from bs4 import BeautifulSoup
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.transaction import atomic
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import get_object_or_404, redirect, get_list_or_404
from django.urls import reverse
from openpyxl.styles import Alignment

from tasks.forms import AddTaskForm, EditTaskForm
from tasks.models import Task, AttachedFile, Engineer, Tag
from tasks.services.service import remove_unused_attached_files


@atomic
def create_tags(request, tags_list):
    """
    Проверяет, существуют ли теги в базе данных, и создаёт новые теги, если они отсутствуют.
    """

    # Получаем данные о тегах из формы (может содержать как новые теги, так и существующие ID)
    tags_data = request.getlist(tags_list)  # Здесь список тегов, включая новые
    new_tag_ids = []  # Для хранения ID тегов (как существующих, так и новых)
    print(tags_data)
    # Получаем все существующие теги для проверки
    existing_tags = {tag.tag_name for tag in Tag.objects.all()}  # Используем множество для быстрого поиска

    existing_lower_tags = {tag.lower() for tag in existing_tags}
    for tag in tags_data:
        tag_lower = tag.lower()
        # Если тег не является числом и не существует в базе, создаем его
        if tag_lower not in existing_lower_tags:
            # Создаем новый тег в базе данных
            new_tag = Tag.objects.create(tag_name=tag)
            # Добавляем ID нового тега в список
            new_tag_ids.append(str(new_tag.id))
        else:
            try:
                # Если тег уже существует, просто добавляем его ID в список
                existing_tag = Tag.objects.get(tag_name=tag)
                new_tag_ids.append(str(existing_tag.id))
                print(existing_tag.id)
            except Exception as e:
                print(f"Тег {tag} уже существует")

    # Копируем данные POST-запроса и заменяем список тегов на список их ID
    post_data = request.copy()
    post_data.setlist(tags_list, new_tag_ids)  # Заменяем теги их ID
    return post_data  # Возвращаем обновленные данные POST-запроса


@login_required
@atomic
def create_task(request):
    redirect_to = reverse("tasks")

    if request.method == "POST":

        post_data = create_tags(request.POST, "tags_create")  # Сохраняем теги и возвращаем

        # Теперь создаем форму с обновлёнными данными (содержит ID всех тегов)
        form = AddTaskForm(post_data, request.FILES, instance=Task(creator=request.user))

        # Получаем URL с параметрами фильтров
        redirect_to = request.POST.get("from_url", redirect_to).strip()

        if form.is_valid():
            task = form.save()

            # Сохраняем прикреплённые файлы (если они есть)
            for file in request.FILES.getlist("files[]"):
                task.files.add(AttachedFile.objects.create(file=file))

            messages.add_message(
                request, messages.SUCCESS, f"Задача '{form.cleaned_data['header']}' успешно создана"
            )
            return redirect(redirect_to)
        else:
            messages.add_message(request, messages.WARNING, form.errors)

    return redirect("tasks")


@login_required
@atomic
def edit_task(request, task_id):
    task = get_object_or_404(Task, pk=task_id)

    # Стандартный редирект на список задач
    redirect_to = reverse("tasks")

    if request.method == "POST":

        post_data = create_tags(request.POST, "tags_edit")  # Сохраняем теги и возвращаем

        form = EditTaskForm(post_data, request.FILES, instance=task)

        # Получаем URL с параметрами фильтров
        redirect_to = request.POST.get("from_url", redirect_to).strip()

        if form.is_valid():
            updated_task = form.save()

            # Удаляем неиспользуемые прикрепленные файлы
            remove_unused_attached_files(request.POST.get("fileuploader-list-files"), updated_task)

            # Обработка прикрепленных файлов
            for file in request.FILES.getlist("files[]"):
                updated_task.files.add(AttachedFile.objects.create(file=file))
            updated_task.save()

            messages.add_message(
                request, messages.SUCCESS, f"Задача '{form.cleaned_data['header']}' отредактирована"
            )
        else:
            messages.add_message(request, messages.WARNING, form.errors)

    # Редирект на исходную страницу с фильтрами
    return redirect(redirect_to)


@login_required
@atomic
def take_task(request, task_id):
    task = get_object_or_404(Task, pk=task_id)
    # Стандартный редирект на список задач
    redirect_to = reverse("tasks")

    if request.method == "POST":
        # Получаем URL с параметрами фильтров
        redirect_to = request.POST.get("from_url", redirect_to).strip()

        try:
            Engineer.objects.get(user=request.user)
            print(request.user.engineer)
            task.engineers.add(request.user.engineer)
            task.save()
            messages.add_message(request, messages.SUCCESS, f"Задача '{task.header}' добавлена в Мои задачи")
        except Engineer.DoesNotExist:
            print(f"у {request.user} нет инженера")
            messages.add_message(request, messages.WARNING, f"у {request.user} нет инженера")
            return redirect(redirect_to)

    # Редирект на исходную страницу
    return redirect(redirect_to)


@login_required
def reopen_task(request, task_id):
    """
    Устанавливает в поле task.is_done = False и добавляет комментарий к полю task.completion_text
    При успешном изменении полей - редирект на страницу откуда была вызвана
    """

    redirect_to: str = reverse("tasks")

    if request.method == "POST":
        task = get_object_or_404(Task, pk=task_id)
        comment = request.POST.get("comment", "")
        # Получаем URL с параметрами фильтров
        redirect_to = request.POST.get("from_url", redirect_to).strip()

        # Обновление задачи
        task.is_done = False
        # task.completion_time = None

        try:
            name = f"{request.user.engineer.first_name} {request.user.engineer.second_name}"
        except AttributeError:
            # Если у пользователя нет engineer, использовать имя пользователя
            name = request.user.username

        pattern = f"\n\nПереоткрыто: [{name} / {datetime.now().strftime('%d.%m.%Y %H:%M')}]\n"

        task.completion_text += pattern + comment
        task.save()
        messages.add_message(request, messages.SUCCESS, f"Задача '{task.header}' возвращена в работу")

    return HttpResponseRedirect(redirect_to)


@login_required
def close_task(request, task_id):
    """
    Устанавливает в поле task.is_done = True и добавляет комментарий к полю task.completion_text
    При успешном изменении полей - редирект на страницу откуда была вызвана
    """
    redirect_to: str = reverse("tasks")

    if request.method == "POST":
        task = get_object_or_404(Task, pk=task_id)
        comment = request.POST.get("comment", "")

        # Получаем URL с параметрами фильтров
        redirect_to = request.POST.get("from_url", redirect_to).strip()

        if "?referrer=" in redirect_to:
            # Извлекаем параметры из URL
            parsed_url = urlparse(redirect_to)
            query_params = parse_qs(parsed_url.query)

            # Получаем параметр referrer
            encoded_referrer = query_params.get("referrer", [""])[0]
            decoded_referrer = unquote(encoded_referrer)

            redirect_to = decoded_referrer

        # Обновление задачи
        task.is_done = True
        # task.completion_time = datetime.now()

        try:
            name = f"{request.user.engineer.first_name} {request.user.engineer.second_name}"
        except AttributeError:
            # Если у пользователя нет engineer, использовать имя пользователя
            name = request.user.username

        pattern = f"\n\nЗакрыто: [{name} / {datetime.now().strftime('%d.%m.%Y %H:%M')}]\n"

        task.completion_text += pattern + comment
        task.save()
        messages.add_message(request, messages.SUCCESS, f"Задача '{task.header}' закрыта")

    return HttpResponseRedirect(redirect_to)


@login_required
def export_to_excel(request):
    tasks_str = request.GET.get("task_ids")
    tasks_list = re.findall(r'\d+', tasks_str)
    tasks_list = list(map(int, tasks_list))

    # Получаем объекты из базы данных
    tasks = get_list_or_404(Task, id__in=tasks_list)

    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    # Установка ширины столбцов и высоты строк
    column_widths = [5, 15, 35, 10, 40, 50, 15, 35, 40, 20, 20, 20, 20]  # Ширина для каждого столбца
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Добавляем заголовки в Excel файл
    headers = ['ID', 'Создатель', "Дата создания", 'Важность', 'Название задачи', 'Описание', 'Задача завершена?',
               'Дата завершения', 'Текст завершения',
               "Инженеры", 'Отделы', 'Теги']
    ws.append(headers)

    # Применение стиля для заголовков (разрешаем перенос строк)
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True,)

    for task in tasks:
        # Получаем значения M2M полей
        engineers = ", ".join([str(engineer) for engineer in task.engineers.all()])
        departments = ", ".join([str(department) for department in task.departments.all()])
        tags = ", ".join([str(tag) for tag in task.tags.all()])

        # Используем BeautifulSoup для извлечения текста из HTML
        soup = BeautifulSoup(task.text, "html.parser")
        text = soup.get_text()

        # Регулярное выражение для удаления изображений, если нужно
        pattern = r'!.*?.*?'  # Находит все строки в формате ![...](...)
        cleaned_text = re.sub(pattern, "", text)

        # Убираем лишние пробелы
        clean_text = re.sub(r'\s+', ' ', cleaned_text).strip()

        # Добавляем строку в Excel и разрешаем перенос строк в ячейках
        row = [
            task.id,
            str(task.creator),
            BeautifulSoup(str(task.create_time), "html.parser").get_text(),
            task.priority,
            task.header,
            clean_text,
            task.is_done,
            BeautifulSoup(str(task.completion_time), "html.parser").get_text(),
            BeautifulSoup(str(task.completion_text), "html.parser").get_text(),
            engineers,
            departments,
            tags,
        ]
        ws.append(row)

        # Применение стиля для каждой ячейки новой строки (включаем перенос текста)
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Возвращаем файл как ответ
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response[
        'Content-Disposition'] = f'attachment; filename=tasks_for_{request.user}_{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}.xlsx'
    wb.save(response)
    return response
