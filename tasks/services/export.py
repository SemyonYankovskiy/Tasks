import re

import openpyxl
from bs4 import BeautifulSoup
from django.db.models import QuerySet
from django.http import HttpResponse
from openpyxl.styles import Alignment

from ..models import Task


class TasksExcelExport:

    def __init__(self, title: str = "tasks"):
        # Создаем Excel файл
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = title
        self._create_header()

    def _create_header(self):
        # Установка ширины столбцов и высоты строк
        column_widths = [5, 15, 35, 10, 40, 50, 15, 35, 40, 20, 20, 20, 20, 40]  # Добавлен новый столбец для объектов
        for i, width in enumerate(column_widths, start=1):
            self.ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Добавляем заголовки в Excel файл
        headers = ['ID', 'Создатель', "Дата создания", 'Важность', 'Название задачи', 'Описание', 'Задача завершена?',
                   'Дата завершения', 'Текст завершения', "Инженеры", 'Отделы', 'Теги', 'Объекты']  # Новый столбец
        self.ws.append(headers)

        # Применение стиля для заголовков (разрешаем перенос строк)
        for cell in self.ws[1]:
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    def add_tasks(self, tasks: QuerySet[Task]):
        for task in tasks:
            # Получаем значения M2M полей
            engineers = ", ".join([str(engineer) for engineer in task.engineers.all()])
            departments = ", ".join([str(department) for department in task.departments.all()])
            tags = ", ".join([str(tag) for tag in task.tags.all()])

            # Получаем объекты, к которым относится задача
            objects = ", ".join([str(obj) for obj in task.objects_set.all()])

            # Используем BeautifulSoup для извлечения текста из HTML
            soup = BeautifulSoup(task.text, "html.parser")
            text = soup.get_text()

            # Регулярное выражение для удаления изображений, если нужно
            pattern = r'!.*?.*?'  # Находит все строки в формате ![...](...)
            cleaned_text = re.sub(pattern, "", text)

            # Убираем лишние пробелы
            clean_text = re.sub(r'\s+', ' ', cleaned_text).strip()

            # Добавляем строку в Excel и разрешаем перенос строк в ячейках
            row = [
                task.id,
                str(task.creator),
                BeautifulSoup(str(task.create_time), "html.parser").get_text(),
                task.priority,
                task.header,
                clean_text,
                task.is_done,
                BeautifulSoup(str(task.completion_time), "html.parser").get_text(),
                BeautifulSoup(str(task.completion_text), "html.parser").get_text(),
                engineers,
                departments,
                tags,
                objects  # Добавляем информацию о связанных объектах
            ]
            self.ws.append(row)

            # Применение стиля для каждой ячейки новой строки (включаем перенос текста)
            for cell in self.ws[self.ws.max_row]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    def make_response(self, filename: str) -> HttpResponse:
        # Возвращаем файл как ответ
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
        self.wb.save(response)
        return response
